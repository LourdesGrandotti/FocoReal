import tkinter as tk
from tkinter import messagebox
import random
import datetime
import os

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False

# ------------------ CONFIG ------------------
MENSAJES = [
    "Volvé al objetivo",
    "Pequeños pasos = grandes logros",
    "No abandones ahora",
    "Tu yo futuro te lo va a agradecer",
    "Concentrate 5 minutos más",
    "La disciplina gana a la motivación"
]

INTERVALO_CHEQUEO = (60000, 90000)  # entre 60 y 90 segundos
INTERVALO_MENSAJES = 5000           # 5 segundos
ARCHIVO_HISTORIAL = "historial_foco.xlsx"

# ------------------ ESTADO ------------------
sesion_activa = False
pausada = False
chequeo_id = None
mensaje_id = None
inicio_sesion = None
total_distracciones = 0
segundos_transcurridos = 0
timer_id = None

# ------------------ FUNCIONES ------------------

def iniciar_sesion():
    global sesion_activa, pausada, inicio_sesion, total_distracciones, segundos_transcurridos

    tarea = tarea_entry.get().strip()
    if not tarea:
        messagebox.showwarning("Atención", "Ingresá una tarea antes de iniciar.")
        return

    sesion_activa = True
    pausada = False
    total_distracciones = 0
    segundos_transcurridos = 0
    inicio_sesion = datetime.datetime.now()

    historial.config(state=tk.NORMAL)
    historial.delete("1.0", tk.END)
    historial.config(state=tk.DISABLED)

    contador_label.config(text="Distracciones: 0")
    btn_iniciar.config(state=tk.DISABLED)
    btn_pausar.config(state=tk.NORMAL, text="Pausar")
    btn_detener.config(state=tk.NORMAL)
    tarea_entry.config(state=tk.DISABLED)

    actualizar_timer()
    programar_chequeo()
    cambiar_mensaje()


def pausar_sesion():
    global pausada, chequeo_id, mensaje_id, timer_id

    if not sesion_activa:
        return

    if not pausada:
        pausada = True
        btn_pausar.config(text="Reanudar")
        if chequeo_id:
            root.after_cancel(chequeo_id)
        if mensaje_id:
            root.after_cancel(mensaje_id)
        if timer_id:
            root.after_cancel(timer_id)
        agregar_historial("--- Sesión pausada ---")
    else:
        pausada = False
        btn_pausar.config(text="Pausar")
        agregar_historial("--- Sesión reanudada ---")
        actualizar_timer()
        programar_chequeo()
        cambiar_mensaje()


def detener_sesion():
    global sesion_activa, chequeo_id, mensaje_id, timer_id

    if not sesion_activa:
        return

    sesion_activa = False
    if chequeo_id:
        root.after_cancel(chequeo_id)
    if mensaje_id:
        root.after_cancel(mensaje_id)
    if timer_id:
        root.after_cancel(timer_id)

    fin = datetime.datetime.now()
    duracion = str(fin - inicio_sesion).split(".")[0]
    agregar_historial(f"--- Sesión finalizada | Duración: {duracion} ---")

    guardar_en_excel(duracion)

    btn_iniciar.config(state=tk.NORMAL)
    btn_pausar.config(state=tk.DISABLED, text="Pausar")
    btn_detener.config(state=tk.DISABLED)
    tarea_entry.config(state=tk.NORMAL)
    timer_label.config(text="00:00:00")
    mensaje_var.set("")


def programar_chequeo():
    global chequeo_id
    if sesion_activa and not pausada:
        proximo = random.randint(*INTERVALO_CHEQUEO)
        chequeo_id = root.after(proximo, chequeo_concentracion)


def chequeo_concentracion():
    if not sesion_activa or pausada:
        return

    ventana = tk.Toplevel(root)
    ventana.title("Chequeo de foco")
    ventana.geometry("300x160")
    ventana.attributes("-topmost", True)
    ventana.resizable(False, False)

    tk.Label(ventana, text="¿Seguís concentrado?", font=("Arial", 13, "bold")).pack(pady=15)

    def concentrado():
        agregar_historial("✔ Concentrado")
        ventana.destroy()
        programar_chequeo()

    def distraido():
        global total_distracciones
        total_distracciones += 1
        contador_label.config(text=f"Distracciones: {total_distracciones}")
        agregar_historial(f"✘ Distracción #{total_distracciones}")
        ventana.destroy()
        programar_chequeo()

    frame_btn = tk.Frame(ventana)
    frame_btn.pack(pady=5)
    tk.Button(frame_btn, text="✔ Sí", width=10, bg="#4CAF50", fg="white",
              font=("Arial", 11), command=concentrado).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_btn, text="✘ No", width=10, bg="#f44336", fg="white",
              font=("Arial", 11), command=distraido).pack(side=tk.LEFT, padx=10)


def cambiar_mensaje():
    global mensaje_id
    if sesion_activa and not pausada:
        mensaje_var.set(random.choice(MENSAJES))
        mensaje_id = root.after(INTERVALO_MENSAJES, cambiar_mensaje)


def actualizar_timer():
    global segundos_transcurridos, timer_id
    if sesion_activa and not pausada:
        segundos_transcurridos += 1
        horas = segundos_transcurridos // 3600
        minutos = (segundos_transcurridos % 3600) // 60
        segundos = segundos_transcurridos % 60
        timer_label.config(text=f"{horas:02d}:{minutos:02d}:{segundos:02d}")
        timer_id = root.after(1000, actualizar_timer)


def agregar_historial(texto):
    hora = datetime.datetime.now().strftime("%H:%M:%S")
    historial.config(state=tk.NORMAL)
    historial.insert(tk.END, f"[{hora}] {texto}\n")
    historial.see(tk.END)
    historial.config(state=tk.DISABLED)


def guardar_en_excel(duracion):
    if not EXCEL_DISPONIBLE:
        return

    tarea = tarea_entry.get().strip()
    fecha = inicio_sesion.strftime("%Y-%m-%d")
    hora_inicio = inicio_sesion.strftime("%H:%M:%S")

    if os.path.exists(ARCHIVO_HISTORIAL):
        wb = load_workbook(ARCHIVO_HISTORIAL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial"
        ws.append(["Fecha", "Hora Inicio", "Tarea", "Duración", "Distracciones"])

    ws.append([fecha, hora_inicio, tarea, duracion, total_distracciones])
    wb.save(ARCHIVO_HISTORIAL)
    agregar_historial(f"📊 Sesión guardada en {ARCHIVO_HISTORIAL}")


# ------------------ INTERFAZ ------------------

root = tk.Tk()
root.title("Control de Foco")
root.geometry("450x520")
root.resizable(False, False)

# Tarea
tk.Label(root, text="Tarea actual:", font=("Arial", 11, "bold")).pack(pady=(15, 2))
tarea_entry = tk.Entry(root, width=45, font=("Arial", 11))
tarea_entry.pack(pady=2)

# Timer
timer_label = tk.Label(root, text="00:00:00", font=("Arial", 28, "bold"), fg="#333")
timer_label.pack(pady=10)

# Botones de control
frame_botones = tk.Frame(root)
frame_botones.pack(pady=5)

btn_iniciar = tk.Button(frame_botones, text="▶ Iniciar", width=12,
                        bg="#4CAF50", fg="white", font=("Arial", 11, "bold"),
                        command=iniciar_sesion)
btn_iniciar.pack(side=tk.LEFT, padx=5)

btn_pausar = tk.Button(frame_botones, text="Pausar", width=12,
                       bg="#FF9800", fg="white", font=("Arial", 11, "bold"),
                       state=tk.DISABLED, command=pausar_sesion)
btn_pausar.pack(side=tk.LEFT, padx=5)

btn_detener = tk.Button(frame_botones, text="■ Detener", width=12,
                        bg="#f44336", fg="white", font=("Arial", 11, "bold"),
                        state=tk.DISABLED, command=detener_sesion)
btn_detener.pack(side=tk.LEFT, padx=5)

# Contador de distracciones
contador_label = tk.Label(root, text="Distracciones: 0", font=("Arial", 11), fg="#555")
contador_label.pack(pady=5)

# Historial
tk.Label(root, text="Historial:", font=("Arial", 10, "bold")).pack()
historial = tk.Text(root, height=8, width=52, state=tk.DISABLED,
                    font=("Courier", 9), bg="#f9f9f9")
historial.pack(pady=5)

# Mensaje motivacional
mensaje_var = tk.StringVar()
mensaje_label = tk.Label(root, textvariable=mensaje_var,
                         font=("Arial", 11, "bold"), fg="#1565C0",
                         wraplength=400)
mensaje_label.pack(pady=10)

if not EXCEL_DISPONIBLE:
    tk.Label(root, text="⚠ openpyxl no instalado: no se guardará en Excel",
             fg="orange", font=("Arial", 9)).pack()

root.mainloop()